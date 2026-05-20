import streamlit as st
import subprocess
import json
import re
import tempfile
import os
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment

# ==========================================
# 辅助函数：执行 Shell 命令并捕获输出
# ==========================================
def run_command(command):
    try:
        result = subprocess.run(
            command, 
            shell=True, 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE, 
            text=True
        )
        return result.returncode, result.stdout, result.stderr
    except Exception as e:
        return -1, "", str(e)

# ==========================================
# 辅助函数：获取当前登录用户信息
# ==========================================
def get_current_user_info():
    """获取当前 Azure 登录的账号信息"""
    code, out, err = run_command("az account show --output json")
    if code == 0:
        try:
            data = json.loads(out)
            return data.get("user", {}).get("name", "Unknown Account")
        except Exception:
            return None
    return None

# ==========================================
# 强制刷新的辅助函数：动态获取订阅
# ==========================================
@st.cache_data(ttl=60) # 设置 60 秒 TTL 缓存，强制定期失效
def get_azure_subscriptions():
    """使用 --refresh 参数获取最新订阅列表"""
    cmd = "az account list --refresh --output json"
    code, out, err = run_command(cmd)
    if code == 0:
        try:
            subs = json.loads(out)
            return [{"name": s["name"], "id": s["id"]} for s in subs]
        except Exception:
            return []
    return []

@st.cache_data
def get_azure_regions():
    """从 Azure 动态获取所有可用区域列表"""
    cmd = "az account list-locations --output json"
    code, out, err = run_command(cmd)
    
    fallback_regions = [
        {"name": "eastus", "displayName": "East US"},
        {"name": "eastus2", "displayName": "East US 2"},
        {"name": "westus", "displayName": "West US"}
    ]
    
    if code == 0:
        try:
            locs = json.loads(out)
            regions = [{"name": l["name"], "displayName": l["displayName"]} for l in locs if "displayName" in l and "name" in l]
            return regions if regions else fallback_regions
        except Exception:
            return fallback_regions
    return fallback_regions

# ==========================================
# 辅助函数：动态获取区域支持的官方模型列表
# ==========================================
@st.cache_data(ttl=300) # 缓存5分钟，避免频繁调用API
def get_official_models(sub_id, location):
    """通过 Azure REST API 获取当前区域支持的所有 Cognitive Services 模型"""
    if not sub_id or not location:
        return []
    
    cmd = f'az rest --method get --uri "https://management.azure.com/subscriptions/{sub_id}/providers/Microsoft.CognitiveServices/locations/{location}/models?api-version=2023-10-01-preview"'
    code, out, err = run_command(cmd)
    
    models_info = []
    if code == 0:
        try:
            data = json.loads(out)
            for item in data.get('value', []):
                # 仅筛选出 OpenAI 相关的模型
                if item.get('model', {}).get('format') == 'OpenAI':
                    model_name = item['model'].get('name', 'Unknown')
                    model_version = item['model'].get('version', 'Unknown')
                    models_info.append({
                        "模型名称 (Name)": model_name,
                        "版本 (Version / Date)": model_version
                    })
            
            # 去重并排序 (优先按模型名称排，再按版本号排倒序)
            unique_models = {f"{m['模型名称 (Name)']}-{m['版本 (Version / Date)']}": m for m in models_info}.values()
            return sorted(list(unique_models), key=lambda x: (x["模型名称 (Name)"], x["版本 (Version / Date)"]), reverse=True)
        except Exception:
            pass
    return []

# ==========================================
# 预设模型列表 (来源于官方文档)
# ==========================================
AVAILABLE_MODELS = [
    "gpt-5.5-2026-04-24",
    "gpt-5.4-2026-03-05",
    "gpt-5.4-pro-2026-03-05",
    "gpt-5.2-2025-12-11",
    "gpt-5.2-pro-2025-12-11",
    "gpt-5.2-chat-2026-02-10",
    "gpt-5.2-chat-2025-12-11",
    "gpt-5.1-2025-11-13",
    "gpt-5-2025-08-07",
    "gpt-5-mini-2025-08-07",
    "gpt-5-chat-2025-08-07",
    "gpt-4.1-2025-04-14",
    "gpt-4o-2024-11-20",
    "gpt-4o-2024-08-06",
    "gpt-4o-2024-05-13",
    "gpt-4-turbo-2024-04-09",
    "o3-2025-04-16",
    "o3-mini-2025-01-31"
]

# ==========================================
# Web 页面布局与逻辑
# ==========================================
st.set_page_config(page_title="Azure 部署助手", page_icon="☁️", layout="wide")

st.title("🚀 Azure OpenAI 自动化部署 Web 工具")
st.markdown("通过左侧面板完成账户登录与参数配置，即可一键执行 Azure 资源的全自动部署。")

# 获取当前登录用户
current_user = get_current_user_info()

# ==========================================
# 左侧边栏：账户管理与参数输入区
# ==========================================
with st.sidebar:
    st.header("🔐 账户管理")
    
    if current_user:
        st.success(f"👤 已登录: **{current_user}**")
        if st.button("🔄 切换账号 / 重新登录", use_container_width=True):
            with st.spinner("正在清理旧账号并等待浏览器登录..."):
                run_command("az account clear")
                code, out, err = run_command("az login --output json")
                if code == 0:
                    get_azure_subscriptions.clear()
                    st.rerun()
                else:
                    st.error("登录被取消或发生错误！")
    else:
        st.error("❌ 当前未登录 Azure")
        if st.button("🔗 点击登录 Azure", type="primary", use_container_width=True):
            with st.spinner("正在清理环境并等待浏览器登录..."):
                run_command("az account clear")
                code, out, err = run_command("az login --output json")
                if code == 0:
                    get_azure_subscriptions.clear()
                    st.rerun()
                else:
                    st.error("登录失败，请检查网络或授权。")
                    
    st.divider()
    st.header("⚙️ 参数配置")
    
    # 动态加载订阅和区域数据
    subscriptions = get_azure_subscriptions() if current_user else []
    regions = get_azure_regions()
    region_options = {r["displayName"]: r["name"] for r in regions}
    region_display_names = list(region_options.keys())
    
    # ---- 订阅选择 ----
    if subscriptions:
        sub_options = {f"{s['name']} ({s['id']})": s['id'] for s in subscriptions}
        selected_sub_label = st.selectbox("选择 Azure 订阅 (Subscription)", options=list(sub_options.keys()))
        selected_sub_id = sub_options[selected_sub_label]
    else:
        if current_user:
            st.error("❌ 未检测到该账号下的有效订阅！")
        selected_sub_id = None
        
    st.divider()
    
    # ---- 区域选择 ----
    rg_default_name = "East US"
    rg_default_index = region_display_names.index(rg_default_name) if rg_default_name in region_display_names else 0
    selected_rg_region_label = st.selectbox("资源组区域 (Resource Group Region)", options=region_display_names, index=rg_default_index)
    selected_rg_region = region_options[selected_rg_region_label]

    openai_default_name = "East US 2"
    openai_default_index = region_display_names.index(openai_default_name) if openai_default_name in region_display_names else 0
    selected_openai_region_label = st.selectbox("OpenAI 资源区域 (OpenAI Region)", options=region_display_names, index=openai_default_index)
    selected_openai_region = region_options[selected_openai_region_label]

    st.divider()
    account_name = st.text_input("账号名 (Account Name)", value="ops", help="例如：ops")
    index_num = st.text_input("序号 (Index)", value="03", help="例如：03")
    
    # ---- 自动生成默认环境名称 ----
    default_rg_name = f"{account_name}-group{index_num}"
    if selected_openai_region == "eastus2":
        suffix = "east-us2"
    elif selected_openai_region == "eastus":
        suffix = "east-us"
    else:
        suffix = selected_openai_region
    default_openai_name = f"group{index_num}-{account_name}-{suffix}"

    st.divider()
    st.markdown("🎯 **1. 核对资源名称 (可直接手动修改)**")
    rg_name = st.text_input("资源组名称", value=default_rg_name)
    openai_name = st.text_input("OpenAI 资源名称", value=default_openai_name)
    
    st.divider()
    st.markdown("👇 **用于第三步：部署模型**")
    
    # 模型快速选择下拉框
    selected_model_full = st.selectbox("从列表中选择预设模型", options=AVAILABLE_MODELS, index=1)
    
    # 自动正则解析
    model_match = re.search(r"^(.*?)-(\d{4}-\d{2}-\d{2}.*)$", selected_model_full)
    if model_match:
        default_model_name = model_match.group(1)
        default_model_version = model_match.group(2)
    else:
        default_model_name = selected_model_full
        default_model_version = ""

    st.markdown("🤖 **2. 核对模型参数 (支持手写新模型)**")
    final_deployment_name = st.text_input("模型部署名称", value=selected_model_full)
    final_model_name = st.text_input("模型真实名称 (Model Name)", value=default_model_name)
    final_model_version = st.text_input("模型真实版本 (Model Version)", value=default_model_version)
    
    # ==========================================
    # 官方当前区域支持的模型列表展示区
    # ==========================================
    st.markdown("---")
    st.markdown("🌐 **官方可用模型查询**")
    if selected_sub_id and selected_openai_region:
        with st.expander(f"查看 `{selected_openai_region}` 区域支持的所有模型"):
            with st.spinner("正在向 Azure 查询..."):
                official_models = get_official_models(selected_sub_id, selected_openai_region)
                if official_models:
                    st.dataframe(official_models, use_container_width=True, hide_index=True)
                    st.caption("提示：在 Azure 中，版本号(Version)通常即为发布日期代号（如 `2024-04-09` 或 `1106-Preview`）。您可以参考上方表格填写模型真实名称与版本。")
                else:
                    st.warning("未能拉取到模型列表。可能是当前订阅在选定区域不支持 OpenAI 资源，或者配额 API 调用受限。")
    
    st.divider()
    filter_name = st.text_input("内容筛选器名称", value="CustomContentFilter123", help="阶段二将以此名称自动创建策略")

# ==========================================
# 主区域：状态预览与提示
# ==========================================
if current_user and selected_sub_id:
    st.info(f"""
    **📝 当前锁定的部署配置预览：**
    * **目标订阅 ID**: `{selected_sub_id}`
    * **资源组名称**: `{rg_name}` (区域: `{selected_rg_region}`)
    * **OpenAI 资源**: `{openai_name}` (区域: `{selected_openai_region}`)
    * **最终部署模型**: `{final_deployment_name}` (👉 真实名称: `{final_model_name}` | 版本: `{final_model_version}`)
    """)
elif not current_user:
    st.warning("👈 请先在左侧边栏登录 Azure 账户。")
elif current_user and not selected_sub_id:
    st.warning("⚠️ 必须配置有效的订阅才能进行部署。")

st.divider()
btn_disabled = (selected_sub_id is None)

# ==========================================
# 操作区：阶段一 (创建基础资源)
# ==========================================
st.subheader("阶段一：创建基础环境")
if st.button("▶️ 1. 一键创建 资源组 & OpenAI 资源", type="primary", disabled=btn_disabled):
    with st.status("正在执行 Azure 部署，请稍候...", expanded=True) as status:
        
        st.write(f"正在配置资源组 `{rg_name}` 到区域 `{selected_rg_region}`...")
        cmd_rg = f"az group create --name {rg_name} --location {selected_rg_region} --subscription {selected_sub_id} --output table"
        code1, out1, err1 = run_command(cmd_rg)
        
        if code1 == 0:
            st.success("资源组创建成功！")
            st.code(out1)
        else:
            status.update(label="部署失败", state="error", expanded=True)
            st.error("资源组创建失败：")
            st.code(err1)
            st.stop() 
            
        st.write(f"正在创建 Azure OpenAI `{openai_name}` 到区域 `{selected_openai_region}`...")
        cmd_openai = f"az cognitiveservices account create --name {openai_name} --resource-group {rg_name} --location {selected_openai_region} --kind OpenAI --sku S0 --custom-domain {openai_name} --subscription {selected_sub_id} --output table"
        code2, out2, err2 = run_command(cmd_openai)
        
        if code2 == 0:
            st.success("Azure OpenAI 资源创建成功！")
            st.code(out2)
            status.update(label="阶段一执行完毕！", state="complete", expanded=True)
            st.balloons()
        else:
            status.update(label="部署失败", state="error", expanded=True)
            st.error("OpenAI 资源创建失败：")
            st.code(err2)

st.divider()

# ==========================================
# 操作区：阶段二 (全自动创建 Guardrails)
# ==========================================
st.subheader("阶段二：自动化配置 Guardrails (内容筛选器)")
st.markdown("通过 API 自动创建最低拦截级别（Lowest blocking / 仅拦截高危风险）的策略。")

if st.button("▶️ 2. 一键创建/更新内容筛选器", type="secondary", disabled=btn_disabled):
    if not filter_name:
        st.warning("⚠️ 请先在左侧边栏填写「内容筛选器名称」！")
    else:
        with st.status(f"正在为 {openai_name} 创建 RAI Policy...", expanded=True) as status:
            
            # 1. 构造 RAI Policy 的 Payload
            rai_payload = {
                "properties": {
                    "mode": "Default",
                    "basePolicyName": "Microsoft.Default",
                    "contentFilters": []
                }
            }
            
            # 遍历四大维度和输入输出双端
            for category in ["hate", "sexual", "violence", "selfharm"]:
                for source in ["Prompt", "Completion"]:
                    rai_payload["properties"]["contentFilters"].append({
                        "name": category,
                        "allowedContentLevel": "High", 
                        "blocking": True, 
                        "enabled": True,
                        "source": source
                    })
            
            # 2. 写入临时文件，避免命令行引号转义噩梦
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json', encoding='utf-8') as f:
                json.dump(rai_payload, f)
                tmp_file_path = f.name
            
            # 3. 使用 az rest 提交 PUT 请求
            st.write(f"正在向 Azure 提交配置: `{filter_name}`...")
            cmd_rai = f'az rest --method put --uri "https://management.azure.com/subscriptions/{selected_sub_id}/resourceGroups/{rg_name}/providers/Microsoft.CognitiveServices/accounts/{openai_name}/raiPolicies/{filter_name}?api-version=2023-10-01-preview" --body @{tmp_file_path}'
            
            code_rai, out_rai, err_rai = run_command(cmd_rai)
            
            # 4. 清理临时文件
            if os.path.exists(tmp_file_path):
                os.remove(tmp_file_path)
            
            if code_rai == 0:
                st.success(f"🎉 内容筛选器 `{filter_name}` 创建成功！四大维度双端均已修改为 Lowest blocking（最低拦截）。")
                status.update(label="阶段二执行完毕！", state="complete", expanded=True)
            else:
                status.update(label="创建失败", state="error", expanded=True)
                st.error("内容筛选器创建失败，可能是当前区域不支持或名称冲突：")
                st.code(err_rai)

st.divider()

# ==========================================
# 操作区：阶段三 (部署模型)
# ==========================================
st.subheader("阶段三：自动化部署模型")
if st.button(f"▶️ 3. 一键部署 {final_deployment_name}", type="primary", disabled=btn_disabled):
    if not filter_name:
        st.warning("请先在左侧填写「内容筛选器名称」！")
    elif not final_deployment_name or not final_model_name:
        st.error("错误：模型部署名称和模型真实名称不能为空！")
    else:
        with st.status(f"正在验证配额并部署 {final_deployment_name}...", expanded=True) as status:
            
            # 探测配额
            target_capacity = 999999
            st.write(f"🔍 正在探测 [{final_model_name}] 模型的绝对最大可用容量...")
            
            cmd_deploy_probe = f"az rest --method put --uri \"https://management.azure.com/subscriptions/{selected_sub_id}/resourceGroups/{rg_name}/providers/Microsoft.CognitiveServices/accounts/{openai_name}/deployments/{final_deployment_name}?api-version=2023-10-01-preview\" --body '{{\"sku\": {{\"name\": \"GlobalStandard\", \"capacity\": {target_capacity}}}, \"properties\": {{\"model\": {{\"format\": \"OpenAI\", \"name\": \"{final_model_name}\", \"version\": \"{final_model_version}\"}}, \"versionUpgradeOption\": \"NoAutoUpgrade\", \"raiPolicyName\": \"{filter_name}\"}}}}'"
            
            code3, out3, err3 = run_command(cmd_deploy_probe)
            
            if code3 != 0 and "InsufficientQuota" in err3:
                match = re.search(r"current available capacity (\d+)", err3)
                if match:
                    max_available = int(match.group(1))
                    
                    if max_available == 0:
                        st.error("🚨 **部署终止！当前剩余可用配额为 0。**\n\n该模型在当前区域的配额已经用光。请先去控制台删除旧模型释放配额。")
                        status.update(label="部署因配额归零而终止", state="error", expanded=True)
                        st.stop()
                    else:
                        st.warning(f"💡 探测成功！当前剩余最大可用容量为 {max_available}。正在为您极限拉满并部署...")
                        
                        # 最终落地部署
                        cmd_deploy_real = f"az rest --method put --uri \"https://management.azure.com/subscriptions/{selected_sub_id}/resourceGroups/{rg_name}/providers/Microsoft.CognitiveServices/accounts/{openai_name}/deployments/{final_deployment_name}?api-version=2023-10-01-preview\" --body '{{\"sku\": {{\"name\": \"GlobalStandard\", \"capacity\": {max_available}}}, \"properties\": {{\"model\": {{\"format\": \"OpenAI\", \"name\": \"{final_model_name}\", \"version\": \"{final_model_version}\"}}, \"versionUpgradeOption\": \"NoAutoUpgrade\", \"raiPolicyName\": \"{filter_name}\"}}}}'"
                        code4, out4, err4 = run_command(cmd_deploy_real)
                        
                        if code4 == 0:
                            st.success(f"🎉 {final_deployment_name} 部署成功！RPM 极限拉满至 {max_available/1000}M，Guardrails 已绑定！")
                            status.update(label="阶段三执行完毕！", state="complete", expanded=True)
                        else:
                            status.update(label="部署失败", state="error", expanded=True)
                            st.error("极限容量部署失败：")
                            st.code(err4)
                else:
                    st.error("无法解析可用容量，原始报错信息：")
                    st.code(err3)
                    status.update(label="探测配额失败", state="error", expanded=True)
            elif code3 == 0:
                st.success(f"🎉 {final_deployment_name} 部署成功！(触发了无限制配额)")
                status.update(label="阶段三执行完毕！", state="complete", expanded=True)
            else:
                status.update(label="部署失败", state="error", expanded=True)
                st.error("探查失败，未知错误：")
                st.code(err3)

st.divider()

# ==========================================
# 操作区：阶段四 (多订阅全自动化导出 Excel 并排版合并合并)
# ==========================================
st.subheader("阶段四：生成并合并多订阅 Excel 配置报表")
st.markdown("点击下方按钮将自动**扫描该账户下的全部订阅**，提取所有 OpenAI 资源的参数，并生成带单元格合并的高级排版表格。")

if st.button("📊 4. 跨全量订阅检索并导出高颜值 Excel 表", type="secondary"):
    if not subscriptions:
        st.error("❌ 当前无有效订阅列表，请先确保左侧登录成功。")
    else:
        with st.status("正在进行跨订阅全量资产扫描与排版（这可能需要一到两分钟）...", expanded=True) as status:
            
            data_rows = []
            
            # 开始遍历当前账号下所有的订阅
            for sub in subscriptions:
                sub_id = sub["id"]
                sub_name = sub["name"]
                
                st.write(f"🔄 正在检索订阅: `{sub_name}` ({sub_id}) ...")
                
                # 1. 检索当前订阅下所有的认知服务账号
                cmd_list = f"az cognitiveservices account list --subscription {sub_id} --output json"
                c_list, o_list, e_list = run_command(cmd_list)
                
                if c_list != 0:
                    st.warning(f"⚠️ 无法读取订阅 `{sub_name}` 的资源列表，已跳过。")
                    continue
                
                try:
                    accounts_json = json.loads(o_list)
                    # 过滤出所有类型为 OpenAI 的资源
                    openai_accounts = [acc for acc in accounts_json if acc.get("kind") == "OpenAI"]
                    
                    if not openai_accounts:
                        st.caption(f"ℹ️ 订阅 `{sub_name}` 内未发现 OpenAI 资源。")
                        continue
                        
                    for acc in openai_accounts:
                        cur_openai_name = acc.get("name")
                        cur_rg_name = acc.get("resourceGroup")
                        cur_region = acc.get("location", "Unknown")
                        cur_endpoint = acc.get("properties", {}).get("endpoint", "获取失败")
                        
                        st.write(f"└─ 🔍 发现 OpenAI 资源: `{cur_openai_name}`，正在拉取密钥与模型...")
                        
                        # 2. 提取当前资源的 API Key
                        cmd_keys = f"az cognitiveservices account keys list --name {cur_openai_name} --resource-group {cur_rg_name} --subscription {sub_id} --output json"
                        c_keys, o_keys, _ = run_command(cmd_keys)
                        cur_api_key = "获取失败"
                        if c_keys == 0:
                            try:
                                cur_api_key = json.loads(o_keys).get("key1", "获取失败")
                            except Exception:
                                pass
                        
                        # 3. 提取当前资源下所有已经部署的模型
                        cmd_deps = f"az cognitiveservices account deployment list --name {cur_openai_name} --resource-group {cur_rg_name} --subscription {sub_id} --output json"
                        c_deps, o_deps, _ = run_command(cmd_deps)
                        
                        cur_deployed_models = []
                        if c_deps == 0:
                            try:
                                deps_json = json.loads(o_deps)
                                cur_deployed_models = [d.get("name", "") for d in deps_json if d.get("name")]
                            except Exception:
                                pass
                        
                        # 如果没有部署模型，则进行保底留痕显示
                        if not cur_deployed_models:
                            cur_deployed_models = ["(未部署任何模型)"]
                        
                        # 4. 按列名组装数据
                        for model_item in cur_deployed_models:
                            data_rows.append({
                                "订阅名": sub_name,
                                "订阅ID": sub_id,
                                "API Key": cur_api_key,
                                "Region": cur_region,
                                "Region Endpoint": cur_endpoint,
                                "Model": model_item
                            })
                            
                except Exception as ex:
                    st.error(f"解析订阅 `{sub_name}` 的 JSON 数据时发生异常: {str(ex)}")
            
            # 判断是否扫描到任何有效数据
            if not data_rows:
                status.update(label="❌ 扫描完成，但未发现任何 OpenAI 资产", state="error", expanded=True)
                st.warning("未能在当前账号的所有订阅中找到任何活动的 Azure OpenAI 资源。")
            else:
                # 5. 转化为 DataFrame 并转换成 Excel 二进制流 (并追加排版逻辑)
                df = pd.DataFrame(data_rows)
                
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='All_Subscriptions_OpenAI')
                    
                    # 获取底层 openpyxl 工作表对象
                    worksheet = writer.sheets['All_Subscriptions_OpenAI']
                    
                    # 设置全局居中样式
                    align_center = Alignment(horizontal='center', vertical='center')
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = align_center
                            
                    # 自适应优雅列宽
                    worksheet.column_dimensions['A'].width = 25  # 订阅名
                    worksheet.column_dimensions['B'].width = 38  # 订阅ID
                    worksheet.column_dimensions['C'].width = 45  # API Key
                    worksheet.column_dimensions['D'].width = 15  # Region
                    worksheet.column_dimensions['E'].width = 65  # Endpoint
                    worksheet.column_dimensions['F'].width = 30  # Model
                    
                    # 核心逻辑：遍历前 5 列进行合并操作（Model列通常不合，保留分行）
                    # openpyxl 列索引是从 1 开始的，所以 1 到 5 就是我们要合并的目标
                    for col_idx in range(1, 6):
                        start_row = 2 # 第一行是表头，数据从第二行开始
                        current_val = worksheet.cell(row=start_row, column=col_idx).value
                        
                        for current_row in range(3, worksheet.max_row + 2):
                            # +2 是为了确保循环到最后一行之后触发最后一次边界判定
                            cell_val = worksheet.cell(row=current_row, column=col_idx).value if current_row <= worksheet.max_row else None
                            
                            # 当内容发生变化，或是已经走到尽头时，执行合并
                            if cell_val != current_val:
                                if current_row - 1 > start_row: # 只有相邻行大于1行时才去合并
                                    worksheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=current_row-1, end_column=col_idx)
                                
                                # 更新指针重新锚定下一个方块
                                start_row = current_row
                                current_val = cell_val
                
                excel_data = output.getvalue()
                status.update(label="🎉 跨订阅扫描及高颜值 Excel 报表生成成功！", state="complete", expanded=False)
                st.success(f"📊 扫描及排版完毕！共计在所有订阅中处理了 **{len(data_rows)} 行** 部署模型数据。")
                
                # 6. 提供统一下载按钮
                st.download_button(
                    label="📥 一键下载精美合并排版 Excel 报表",
                    data=excel_data,
                    file_name="Azure_OpenAI_Inventory_Formatted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
